from textfsm import TextFSM
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl
from datetime import datetime
import wcwidth


class SP2EXCEL:

    # 初始化一些内容
    def __init__(self, echo_content):
        self.sp_textfsm_path = 'hp_comware_display_security-policy_ip_local.textfsm'
        self.obj_ip_textfsm_path = 'hp_comware_display_object-group_ip_local.textfsm'
        self.obj_service_textfsm_path = 'hp_comware_display_object-group_service_local.textfsm'

        # 回显内容读入内存
        with open(echo_content, 'r', encoding='utf8') as f:
            self.contents = f.read()


    # 0、字典的value是列表的，将列表拼接为一个以换行符为分隔符的字符串
    def expand_list(self, list_dict_res):
        for d in list_dict_res:
            for k, v in d.copy().items():
                if isinstance(v, list):
                    # 将列表拼接为一个以换行符为分隔符的字符串
                    obj_item_str = '\n'.join(v)
                    # 写回原字典
                    d[k] = obj_item_str

        return list_dict_res


    # 1、解析 dis security-policy ip 回显内容（安全策略）
    def parse_sp(self):

        with open(self.sp_textfsm_path, encoding='utf8') as f:
            template = TextFSM(f)
            res = template.ParseTextToDicts(self.contents)

            # 对没有匹配到内容的设置默认值
            for d in res:
                if d['ACTIVE_STATE'] == '':
                    d['ACTIVE_STATE'] = 'Active'
                if d['VRF'] == '':
                    d['VRF'] = 'public'
                if d['PROFILE'] == '':
                    d['PROFILE'] = 'none'
                if d['LOGGING'] == '':
                    d['LOGGING'] = 'disable'
                if not d['COUNTING']:
                    d['COUNTING'] = 'disable'
                if d['TIME_RANGE'] == '':
                    d['TIME_RANGE'] = 'none'
                if d['DESC'] == '':
                    d['DESC'] = 'none'
                if not d['SESSION']:
                    d['SESSION'] = 'none'
                if not d['SRC_ZONE']:
                    d['SRC_ZONE'] = 'any'
                if not d['DEST_ZONE']:
                    d['DEST_ZONE'] = 'any'
                if not d['SRC_IP']:
                    d['SRC_IP'] = 'any'
                if not d['DEST_IP']:
                    d['DEST_IP'] = 'any'
                if not d['SERVICE']:
                    d['SERVICE'] = 'any'
                if not d['APPLICATION']:
                    d['APPLICATION'] = 'any'
                if not d['USER']:
                    d['USER'] = 'any'

            # 调用expand_list()处理字典value为列表的键值对
            res = self.expand_list(res)

            df = pd.DataFrame(res)

            # 安全策略表头
            df_th = {'RULE_ID': 'ID', 'RULE_NAME': '名称', 'ACTIVE_STATE': '时间段状态', 'ACTION': '动作', 'VRF': '公网', 'PROFILE': '内容安全', 'LOGGING': '日志', 'COUNTING': '统计', 'TIME_RANGE': '时间段', 'DESC': '描述', 'SESSION': '会话', 'SRC_ZONE': '源安全域', 'DEST_ZONE': '目的安全域', 'SRC_IP': '源地址', 'DEST_IP': '目的地址', 'SERVICE': '服务', 'APPLICATION': '应用', 'USER': '用户'}

            self.sp_df = df.rename(columns=df_th)

            # 下一步进行解析地址对象组回显内容
            self.parse_obj_ip()


    # 2、解析 dis object-group ip address  回显内容（地址对象组）
    def parse_obj_ip(self):

        with open(self.obj_ip_textfsm_path, encoding='utf8') as f:
            template = TextFSM(f)
            res = template.ParseTextToDicts(self.contents)

            # 调用expand_list()处理字典value为列表的键值对
            res = self.expand_list(res)

            df = pd.DataFrame(res)

            # 地址对象组表头
            df_th = {'OBJ_GROUP_NAME': '对象组名称', 'OBJ_ITEM': '对象', 'REFERENCED': '被引用', 'SEC_ZONE': '安全域', 'DESC': '描述'}
            self.obj_ip_df = df.rename(columns=df_th)

            # 下一步进行解析服务对象组回显内容
            self.parse_obj_service()

    # 3、解析 dis object-group service  回显内容（服务对象组）
    def parse_obj_service(self):

        with open(self.obj_service_textfsm_path, encoding='utf8') as f:
            template = TextFSM(f)
            res = template.ParseTextToDicts(self.contents)

            # 调用expand_list()处理字典value为列表的键值对
            res = self.expand_list(res)

            df = pd.DataFrame(res)

            # 服务对象组表头
            df_th = {'OBJ_GROUP_NAME': '对象组名称', 'OBJ_ITEM': '对象', 'REFERENCED': '被引用', 'DESC': '描述'}
            self.obj_service_df = df.rename(columns=df_th)
            

    # 4、表格写入
    def to_excel(self):

        # 运行以上三个解析内容的函数，一个串一个
        self.parse_sp()

        # 文件名时间部分，用于区别新旧
        create_time = datetime.now().strftime("%Y%m%d%H%M%S")
        self.output_path = f'核心防火墙安全策略统计{create_time}.xlsx'

        # 创建ExcelWriter对象并指定文件名
        with pd.ExcelWriter(self.output_path) as writer:
            # 将三个DataFrame写入不同的工作表中，并且不写入行索引
            self.sp_df.to_excel(writer, sheet_name='安全策略', index=False)
            self.obj_ip_df.to_excel(writer, sheet_name='地址对象组', index=False)
            self.obj_service_df.to_excel(writer, sheet_name='服务对象组', index=False)

            # 取出wb对象供define_excel_style()使用
            self.wb = writer.book
            self.sheet_names = self.wb.sheetnames

        self.define_excel_style()


    # 5、调整表格样式
    def define_excel_style(self):

        # 三个sheet
        self.ws1 = self.wb[self.sheet_names[0]]
        self.ws2 = self.wb[self.sheet_names[1]]
        self.ws3 = self.wb[self.sheet_names[2]]

        # 以单元格中当行最长字符串长度调整列宽
        for sheet_name in self.sheet_names:
            worksheet = self.wb[sheet_name]
            # 遍历所有列，找出最宽的一列并调整列宽
            for col in worksheet.columns:
                max_width = 0
                column = col[0].column # openpyxl 中的列索引从 1 开始
                
                # 计算每个单元格中最宽行的显示宽度，并选取最宽的一行
                for cell in col:
                    lines = str(cell.value).split("\n")
                    max_line_width = 0
                    for line in lines:
                        width = wcwidth.wcswidth(line) + 2
                        if width > max_line_width:
                            max_line_width = width
                    if max_line_width > max_width:
                        max_width = max_line_width
                
                # 将最宽行的显示宽度设置为该列的列宽
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = max_width

        # 冻结首行
        self.ws1.freeze_panes = 'A2'
        self.ws2.freeze_panes = 'A2'
        self.ws3.freeze_panes = 'A2'

        # 表头背景色
        th_color_fill = PatternFill('solid', fgColor='D3D3D3')
        for col in range(1, 19):
            self.ws1.cell(row=1, column=col).fill = th_color_fill
        for col in range(1, 6):
            self.ws2.cell(row=1, column=col).fill = th_color_fill
        for col in range(1, 6):
            self.ws3.cell(row=1, column=col).fill = th_color_fill

        # 垂直居中
        for sheet_name in self.sheet_names:
            worksheet = self.wb[sheet_name]
            for r in worksheet:
                for c in r:
                    c.alignment = openpyxl.styles.Alignment(vertical='center',
                                                        wrapText=True)

        # 保存
        self.wb.save(self.output_path)


if __name__ == '__main__':

    # 指定回显输出文件路径，同目录下直接 文件名
    echo_content_path = 'output.log'
    # 实例化对象
    obj = SP2EXCEL(echo_content_path)
    # 执行 4、表格写入
    obj.to_excel()

    print('搞快点-->>搞快点-->>')
